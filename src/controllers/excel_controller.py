import polars as pl
from pathlib import Path

from src.helpers import excel_helpers
from src.config import REQUIRED_COLUMNS, TICKET_COLUMNS


from pathlib import Path
import polars as pl

from src.helpers.datetime_helpers import normalize_fecha_hora_polars, split_web_creation_dt

from openpyxl import load_workbook
from src.models.ticket_job import TicketJob


class ExcelController:
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.df = None

        self.format = None
        self.ticket_column = None

        self._headers = None
        self._header_row_df = None

        self._run()

    def _run(self):
        self._validate_file()
        self._load_excel()
        # self._validate_structure()
        # self._filter_pending()

    def _validate_file(self):
        if not self.excel_path.exists():
            raise FileNotFoundError("El archivo Excel no existe")
        
    def _load_excel(self):
        df_raw = excel_helpers.read_excel_with_excel_row(self.excel_path)

        if df_raw.is_empty():
            raise ValueError("El archivo Excel no contiene datos")
        
        header_row = excel_helpers.detect_header_row(df_raw)
        raw_headers = list(df_raw.row(header_row))[1:]
        headers = excel_helpers.clean_headers([str(h) for h in raw_headers])

        self._headers = headers
        self._header_row_df = header_row

        self.format = excel_helpers.detect_format(headers)
        self.ticket_column = excel_helpers.validate_required_columns(headers, REQUIRED_COLUMNS, TICKET_COLUMNS)
        
        df_data = df_raw.slice(header_row + 1)
        if self.format == "NEW":
            df_data = df_data.slice(1)

        df_data = df_data.select(["EXCEL_ROW"] + df_data.columns[1:1 + len(headers)])
        df_data.columns = ["EXCEL_ROW"] + headers
        df_data = df_data.filter(pl.any_horizontal(pl.all().is_not_null()))
        df_data = normalize_fecha_hora_polars(df_data)
        df_data = excel_helpers.reduce_to_core_columns(df = df_data, ticket_col= self.ticket_column)
        df_data = excel_helpers.filter_pending_tickets(df=df_data, ticket_col="TICKET")

        if df_data.is_empty():
            raise ValueError("La planilla no contiene ticket pendientes para cargar")

        self.df = df_data

        print("df_data")
        print(df_data)

        self.df.write_csv("debug_output.csv")

    def _excel_col_index(self, header_name: str) -> int:
        # OJO: headers corresponden a columnas reales del Excel empezando desde col_2
        hu = [h.strip().upper() for h in self._headers]
        idx = hu.index(header_name.strip().upper())  # 0-based en headers
        return idx + 2  # +2 => columna Excel real (B=2)


    def add_datetime(self, job):
        if not job.creation_dt_text:
            return

        web_date, _ = split_web_creation_dt(job.creation_dt_text)

        wb = load_workbook(self.excel_path)
        ws = wb.worksheets[0]

        r = int(job.row_id)
        c = self._excel_col_index("FECHA")
        cell = ws.cell(row=r, column=c)

        if cell.value in (None, "", "NONE"):
            print(f"📅 Agregando FECHA en Excel fila {r}")
            cell.value = web_date

            if cell.number_format in (None, "", "General"):
                cell.number_format = "dd-mm-yyyy"  # o "dd/mm/yyyy"

        wb.save(self.excel_path)


    def add_time(self, job):
        if not job.creation_dt_text:
            return

        _, web_time = split_web_creation_dt(job.creation_dt_text)

        wb = load_workbook(self.excel_path)
        ws = wb.worksheets[0]

        r = int(job.row_id)
        c = self._excel_col_index("HORA")
        cell = ws.cell(row=r, column=c)

        if cell.value in (None, "", "NONE"):
            print(f"🕒 Agregando HORA en Excel fila {r}")
            cell.value = web_time

            # Mantiene fill/border/font/alignment (destino).
            # Ajusta SOLO el formato numérico si está en General:
            if cell.number_format in (None, "", "General"):
                cell.number_format = "hh:mm"  # o "hh:mm:ss"

        wb.save(self.excel_path)


    def add_ticket(self, job: TicketJob):
        print(f"✍️ Registrando ticket en Excel fila {job.row_id}")


    def return_excel(self):
        pass