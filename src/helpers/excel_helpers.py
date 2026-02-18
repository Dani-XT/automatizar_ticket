import polars as pl
from typing import List, Tuple, Set
from openpyxl import load_workbook
from datetime import datetime, date, time as dtime
from pathlib import Path
from src.config import CORE_COLUMNS

from src.models.ticket_job import TicketJob

from src.helpers.datetime_helpers import split_web_creation_dt


def clean_headers(headers: List[str]) -> List[str]:
    return [
        h.strip()
        for h in headers
        if h and str(h).strip().upper() != "NONE"
    ]

def detect_header_row(df_raw: pl.DataFrame) -> int:
    for i, row in enumerate(df_raw.iter_rows()):
        normalized = {
            str(c).strip().upper()
            for c in row
            if c
        }
        if "FECHA" in normalized and "HORA" in normalized:
            return i
    
    raise ValueError("No se pudo detectar la fila de encabezados")

def detect_format(headers: List[str]) -> str:
    normalized = {h.upper() for h in headers}

    if "TKT" in normalized and "TICKET" not in normalized:
        return "OLD"
    
    if "TICKET" in normalized and "EDIFICIO" in normalized:
        return "NEW"
    
    raise ValueError("Formato de planilla no reconocido")

def validate_required_columns( headers: List[str], required: Set[str], ticket_aliases: Set[str]) -> str:
    normalized = {h.upper(): h for h in headers}
    missing = [
        col for col in required
        if col not in normalized
    ]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")
    
    for alias in ticket_aliases:
        if alias in normalized:
            return normalized[alias]
        
    raise ValueError("No se encontro columna de ticket (TKT o Ticket)")




def reduce_to_core_columns(df: pl.DataFrame, ticket_col: str) -> pl.DataFrame:

    # Normalizar nombre del ticket
    if ticket_col != "TICKET":
        df = df.rename({ticket_col: "TICKET"})

    missing = [c for c in CORE_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Columnas críticas faltantes: {', '.join(missing)}")
    
    cols = (["EXCEL_ROW"] if "EXCEL_ROW" in df.columns else []) + CORE_COLUMNS
    return df.select(cols)

def filter_pending_tickets(df: pl.DataFrame, ticket_col: str) -> pl.DataFrame:
    return df.filter(
        pl.col(ticket_col).is_null()
        | (pl.col(ticket_col)
           .cast(pl.Utf8, strict=False)
           .str.strip_chars() == "")
        | (pl.col(ticket_col)
           .cast(pl.Utf8, strict=False)
           .str.to_uppercase() == "NONE")
    )

def read_excel_with_excel_row(path: Path, sheet_name: str | None = None) -> pl.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    max_row, max_col = ws.max_row, ws.max_column

    used_cols = [False] * max_col   # marca columnas que tienen algún dato real
    rows = []

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True),
        start=1
    ):
        if all(v is None or v == "" for v in row):
            continue

        vals = []
        for j, v in enumerate(row, start=1):
            if v is None or v == "":
                vals.append(None)
                continue

            # hay valor real => esta columna se usa
            used_cols[j - 1] = True

            if isinstance(v, (datetime, date, dtime)):
                vals.append(v.isoformat())
            else:
                s = str(v).strip()
                vals.append(s if s != "" else None)
                # si era solo espacios, no cuenta como dato real
                if s == "":
                    used_cols[j - 1] = False

        rows.append([r_idx, *vals])

    keep = [i + 1 for i, ok in enumerate(used_cols) if ok]
    cropped = []
    for r in rows:
        excel_row = r[0]
        data = r[1:]
        cropped.append([excel_row] + [data[i - 1] for i in keep])

    schema = ["EXCEL_ROW"] + [f"col_{i}" for i in keep]
    return pl.DataFrame(cropped, schema=schema, orient="row")





